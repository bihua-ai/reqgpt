from py2neo import Graph
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
import oms_library_gdb, oms_library_project
import seagent_file
import string
import os, json, seagent_file
import streamlit as st
from dotenv import load_dotenv

st.session_state.testset = None

border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

bold_font = Font(bold=True)
# Colors and borders setting:
odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
heading_row_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    # Apply styles to the cell
    # cell = ws['A1']
    # cell.fill = fill
    # cell.border = border

def set_cell_content(wb, row, column, value):
    cell = wb.cell(row, column, value)
    cell.border = border
    if row % 2 != 0: # odd
        cell.fill = odd_row_fill
    else:
        cell.fill = even_row_fill
    cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')
    


# Initialize the Neo4j graph connection

# def oms_library_gdb.cypher_execution(query, parameters=None):
#     graph = Graph("bolt://localhost:7687", auth=("neo4j", "sw2201!@#"))
#     return graph.run(query, parameters)

def get_list_of_start_nodes(project_name, testcase_scope):
    query = """
    MATCH (o:car {project_name: $project_name})
    WHERE NOT ()-[:NEXT]->(o)
    RETURN o.identifier AS identifier
    """
    result = oms_library_gdb.cypher_execution(query, {"project_name": project_name})
    records = [record["identifier"] for record in result]
    print(f"get_list_of_end_nodes = {records}")
    return records

def get_list_of_end_nodes(project_name, testcase_scope):
    query = """
    MATCH (o:car {project_name: $project_name})
    WHERE NOT (o)-[:NEXT]->()
    RETURN o.identifier AS identifier
    """
    result = oms_library_gdb.cypher_execution(query, {"project_name": project_name})
    records = [record["identifier"] for record in result]
    print(f"get_list_of_end_nodes = {records}")
    return records
    # return [record["identifier"] for record in result]

def add_common_upstream(project_name):
    query_create = """
    CREATE (commonUpstream:car {identifier: 'common_upstream', project_name: $project_name})
    """
    oms_library_gdb.cypher_execution(query_create, {"project_name": project_name})
    
    query_link = """
    MATCH (o:car {project_name: $project_name})
    WHERE NOT (o)<-[:NEXT]-()
    MATCH (commonUpstream:car {identifier: 'common_upstream', project_name: $project_name})
    CREATE (commonUpstream)-[:NEXT]->(o)
    """
    oms_library_gdb.cypher_execution(query_link, {"project_name": project_name})

    query_delete = """
    MATCH (commonUpstream:car {identifier: 'common_upstream', project_name: $project_name})-[r:NEXT]-(common_upstream:car {identifier: 'common_upstream', project_name: $project_name})
    DELETE r
    """
    oms_library_gdb.cypher_execution(query_delete, {"project_name": project_name})

def add_common_downstream(project_name):
    query_create = """
    CREATE (commonDownstream:car {identifier: 'common_downstream', project_name: $project_name})
    """
    oms_library_gdb.cypher_execution(query_create, {"project_name": project_name})
    
    query_link = """
    MATCH (o:car {project_name: $project_name})
    WHERE NOT (o)-[:NEXT]->()
    MATCH (commonDownstream:car {identifier: 'common_downstream', project_name: $project_name})
    CREATE (o)-[:NEXT]->(commonDownstream)
    """
    oms_library_gdb.cypher_execution(query_link, {"project_name": project_name})

    query_delete = """
    MATCH (commonDownstream:car {identifier: 'common_downstream', project_name: $project_name})-[r:NEXT]-(commonDownstream:car {identifier: 'common_downstream', project_name: $project_name})
    DELETE r
    """
    oms_library_gdb.cypher_execution(query_delete, {"project_name": project_name})

def remove_common_nodes():
    query_upstream = """
    MATCH (commonUpstream:car {identifier: 'common_upstream'})
    DETACH DELETE commonUpstream
    """
    oms_library_gdb.cypher_execution(query_upstream)
    
    query_downstream = """
    MATCH (commonDownstream:car {identifier: 'common_downstream'})
    DETACH DELETE commonDownstream
    """
    oms_library_gdb.cypher_execution(query_downstream)

def get_all_paths_as_json(start_identifier, end_identifier, project_name):
    query = """
    MATCH p = (start:car {identifier: $start_identifier, project_name: $project_name})-[:NEXT*]->(end:car {identifier: $end_identifier, project_name: $project_name})
    RETURN p
    """
    result = oms_library_gdb.cypher_execution(query, {"start_identifier": start_identifier, "end_identifier": end_identifier, "project_name": project_name})
    print(result)

    paths = []
    for record in result:
        path = record["p"]
        nodes = [node["identifier"] for node in path.nodes]
        paths.append(nodes)
    return paths
    # return json.dumps({"paths": paths}, indent=4)

def convert_to_testset_json_from_gdb_output(project_name, paths_json, spec_json_home):
    # build car hashtable
    car_hashtable = {}
    variable_hashtable = {}
    attribute_eqc_hasttable = {} # identifier - name or description
    state_hashtable = {}

    # load_dotenv()
    # SPEC_SUB_DIRECTORY = os.getenv("SPEC_SUB_DIRECTORY")
    # spec_json_home = os.path.join(project_path, SPEC_SUB_DIRECTORY)
    system_under_test = None 
    for file_entry in os.scandir(spec_json_home):
        # Check if the entry is a file. It is always a file, there is no subdirectory here.
        if file_entry.is_file():
            spec_path = os.path.join(spec_json_home, file_entry)
            spec_json = seagent_file.oms_load(spec_path)
            user_id = st.session_state.app_state["user_identifier"]
            suffix = f"_{user_id}"
            system_under_test = project_name[:-len(suffix)]
            # system_under_test = system_under_test or spec_json["info"]["title"]

            for i in range(len(spec_json["omsObject"]["memberObjects"])):

                # variable part
                variable_identifier = spec_json["omsObject"]["memberObjects"][i]["identifier"]
                variable_hashtable[variable_identifier] = spec_json["omsObject"]["memberObjects"][i]

                # attribute part
                number_of_attributes = len(spec_json["omsObject"]["memberObjects"][i]["attributes"])
                for j in range(number_of_attributes):
                    attribute_identifier = spec_json["omsObject"]["memberObjects"][i]["attributes"][j]["identifier"]
                    attribute_eqc_hasttable[attribute_identifier] = spec_json["omsObject"]["memberObjects"][i]["attributes"][j]

                    # equivalence part
                    number_of_equivalence_classes = len(spec_json["omsObject"]["memberObjects"][i]["attributes"][j]["equivalenceClasses"])
                    for k in range(number_of_equivalence_classes):
                        eqc_identifier = spec_json["omsObject"]["memberObjects"][i]["attributes"][j]["equivalenceClasses"][k]["identifier"]
                        attribute_eqc_hasttable[eqc_identifier] = spec_json["omsObject"]["memberObjects"][i]["attributes"][j]["equivalenceClasses"][k]

                # state part
                number_of_states = len(spec_json["omsObject"]["memberObjects"][i]["states"])
                for l in range(number_of_states):
                    state_identifier = spec_json["omsObject"]["memberObjects"][i]["states"][l]["identifier"]
                    state_element = {}
                    state_element["variable_identifier"] = spec_json["omsObject"]["memberObjects"][i]["identifier"]
                    state_element["state_object"] = spec_json["omsObject"]["memberObjects"][i]["states"][l]
                    state_hashtable[state_identifier] = state_element

            # car part
            for i in range(len(spec_json["omsObject"]["actions"])):
                number_of_cars = len(spec_json["omsObject"]["actions"][i]["cars"])
                for j in range(number_of_cars):
                    car_identifier = spec_json["omsObject"]["actions"][i]["cars"][j]["identifier"]
                    car_element = {}
                    car_element["action_name"] = spec_json["omsObject"]["actions"][i]["name"]
                    car_element["car"] = spec_json["omsObject"]["actions"][i]["cars"][j]
                    car_hashtable[car_identifier] = car_element

    # seagent_file.oms_save("./car_hashtable.json", car_hashtable)

    # construct testset
    testset_json = {}
    testset_json["system_under_test"] = system_under_test
    testset_json["module_under_test"] = system_under_test
    testset_json["test_cases"] = []

    # print("convert_to_testset_json_from_gdb_output")
    # print(paths_json)
    ###########################################################################
    number_of_testcases = len(paths_json)
    for i in range(number_of_testcases):
        testcase = {}
        testcase["test_case_name"] = f"测试用例 {i + 1}"
        testcase["test_steps"] = []

        number_of_steps = len(paths_json[i])
        # skip first (common_upstream) and last (common_downstream)
        # real number of steps = number_of_steps - 2
        for j in range(number_of_steps - 1): 
            if  paths_json[i][j] == "common_upstream":
                continue # skip the first fake element
            teststep = {}

            # step general
            teststep["test_step_number"] = str(j)
            car_identifier = paths_json[i][j]
            car_element = car_hashtable[car_identifier]
            teststep["car_name"] = car_element["car"]["name"]

            # condition of car
            teststep["condition"] = {}
            teststep["condition"]["condition_description"] = car_element["car"]["condition"]["description"]
            teststep["condition"]["condition_states"] = []
            number_of_states = len(car_element["car"]["condition"]["states"])
            for k in range(number_of_states): 
                state_identifier = car_element["car"]["condition"]["states"][k]   
                state_element = state_hashtable[state_identifier]
                variable_identifier = state_element["variable_identifier"]
                variable_element = variable_hashtable[variable_identifier]

                condition_state_element = {}
                condition_state_element["variable_name"] = variable_element["name"]
                condition_state_element["state_name"] = state_element["state_object"]["name"]
                condition_state_element["attribute_equivalence_classes"] = []
        
                # teststep["condition"]["condition_states"][k]["variable_name"] = variable_element["name"]
                # teststep["condition"]["condition_states"][k]["state_name"] = state_element["state_object"]["name"]

                # teststep["condition"]["condition_states"][k]["attribute_equivalence_classes"] = []
                
                number_of_attribute_equivalence_classes = len(state_element["state_object"]["attributeAndEquivalenceClasses"])
                for l in range(number_of_attribute_equivalence_classes):
                    
                    attributeId = state_element["state_object"]["attributeAndEquivalenceClasses"][l]["attributeId"]
                    attribute_element = attribute_eqc_hasttable[attributeId]
                    attribute_name = attribute_element["name"]
                    attribute_description = attribute_element["description"]

                    equivalenceClassId = state_element["state_object"]["attributeAndEquivalenceClasses"][l]["equivalenceClassId"]
                    equivalence_class_element = attribute_eqc_hasttable[equivalenceClassId]
                    equivalence_class_name = equivalence_class_element["name"]
                    equivalence_class_description = equivalence_class_element["description"]

                    attribute_equivalence_classe_element = {}
                    attribute_equivalence_classe_element["attribute_name"] = attribute_name
                    attribute_equivalence_classe_element["attribute_description"] = attribute_description
                    attribute_equivalence_classe_element["equivalence_class_name"] = equivalence_class_name
                    attribute_equivalence_classe_element["equivalence_class_description"] = equivalence_class_description

                    condition_state_element["attribute_equivalence_classes"].append(attribute_equivalence_classe_element)

                    # teststep["condition"]["condition_states"][k]["attribute_equivalence_classes"][l]["attribute_name"] = attribute_name
                    # teststep["condition"]["condition_states"][k]["attribute_equivalence_classes"][l]["attribute_description"] = attribute_description
                    # teststep["condition"]["condition_states"][k]["attribute_equivalence_classes"][l]["equivalence_class_name"] = equivalence_class_name
                    # teststep["condition"]["condition_states"][k]["attribute_equivalence_classes"][l]["equivalence_class_description"] = equivalence_class_description   
                teststep["condition"]["condition_states"].append(condition_state_element)           

            # action of car
            teststep["action_name"] = car_element["action_name"]

            # result of car
            teststep["result"] = {}
            teststep["result"]["result_description"] = car_element["car"]["result"]["description"]
            teststep["result"]["result_states"] = []
            number_of_states = len(car_element["car"]["result"]["states"])
            for k in range(number_of_states): 
                state_identifier = car_element["car"]["result"]["states"][k]   
                state_element = state_hashtable[state_identifier]
                variable_identifier = state_element["variable_identifier"]
                variable_element = variable_hashtable[variable_identifier]

                result_state_element = {}
                result_state_element["variable_name"] = variable_element["name"]
                result_state_element["state_name"] = state_element["state_object"]["name"]
                result_state_element["attribute_equivalence_classes"] = []
        
                # teststep["result"]["result_states"][k]["variable_name"] = variable_element["name"]
                # teststep["result"]["resultn_states"][k]["state_name"] = state_element["state_object"]["name"]
                # teststep["result"]["condition_states"][k]["attribute_equivalence_classes"] = []

                number_of_attribute_equivalence_classes = len(state_element["state_object"]["attributeAndEquivalenceClasses"])
                for l in range(number_of_attribute_equivalence_classes):
                    attributeId = state_element["state_object"]["attributeAndEquivalenceClasses"][l]["attributeId"]
                    attribute_element = attribute_eqc_hasttable[attributeId]
                    attribute_name = attribute_element["name"]
                    attribute_description = attribute_element["description"]

                    equivalenceClassId = state_element["state_object"]["attributeAndEquivalenceClasses"][l]["equivalenceClassId"]
                    equivalence_class_element = attribute_eqc_hasttable[equivalenceClassId]
                    equivalence_class_name = equivalence_class_element["name"]
                    equivalence_class_description = equivalence_class_element["description"]

                    attribute_equivalence_classe_element = {}
                    attribute_equivalence_classe_element["attribute_name"] = attribute_name
                    attribute_equivalence_classe_element["attribute_description"] = attribute_description
                    attribute_equivalence_classe_element["equivalence_class_name"] = equivalence_class_name
                    attribute_equivalence_classe_element["equivalence_class_description"] = equivalence_class_description

                    result_state_element["attribute_equivalence_classes"].append(attribute_equivalence_classe_element)

                    # teststep["condition"]["condition_states"][k]["attribute_equivalence_classes"][l]["attribute_name"] = attribute_name
                    # teststep["condition"]["condition_states"][k]["attribute_equivalence_classes"][l]["attribute_description"] = attribute_description
                    # teststep["condition"]["condition_states"][k]["attribute_equivalence_classes"][l]["equivalence_class_name"] = equivalence_class_name
                    # teststep["condition"]["condition_states"][k]["attribute_equivalence_classes"][l]["equivalence_class_description"] = equivalence_class_description    
                teststep["result"]["result_states"].append(result_state_element) 
            testcase["test_steps"].append(teststep)
        testset_json["test_cases"].append(testcase)
        # seagent_file.oms_save("./tmp.json", testset_json)
    return testset_json

import time

# from neo4j database to json
def generate_testset_json_from_gdb(specification_directory, project_name, testcase_scope):
    # project_name = '安全文件上传与登录验证系统'
    start_identifier = 'common_upstream'
    end_identifier = 'common_downstream'
    testset_json = None
    try:
        timestamp = time.time()
        print("开始（秒）:", timestamp)
        start_nodes = get_list_of_start_nodes(project_name, testcase_scope)
        end_nodes = get_list_of_end_nodes(project_name, testcase_scope)
        print("Start Nodes:", start_nodes)
        print("End Nodes:", end_nodes)
        
        add_common_upstream(project_name)
        add_common_downstream(project_name)
        
        paths = get_all_paths_as_json(start_identifier, end_identifier, project_name)
        print("Paths JSON:", paths)
        # seagent_file.oms_save("./test_paths.json", paths)
        
        # Remove the common nodes after path retrieval
        remove_common_nodes()

        # clear db so db can be light and ready for other to use.
        oms_library_gdb.remove_project_from_gdb(project_name)

        # convert to testset_json
        testset_json = convert_to_testset_json_from_gdb_output(project_name, paths, specification_directory)
        # timestamp = time.time()
        # print("变成json完成（秒）:", timestamp)

        return testset_json

    except Exception as e:
        print("An error occurred:", e)
    finally:
        return testset_json

def testset_convert_json_to_xlsx(testset_json):
    # seagent_file.oms_save("./testset_json.json", testset_json)
    # path = "/opt/bihua/reqgpt/seagent/testcase_template.json"
    # testset_json = seagent_file.oms_load(path)

    wb = Workbook()
    ws = wb.active
    ws.title = testset_json["system_under_test"]
    
    # current_row_number = 1
    # headers = [
    # "用例名称", "测试步骤", "条件概述", "条件细节", "行动", "结果概述", "结果细节", "测试结果", "复测结果", "备注"
    # ]
    headers = [
    "用例名称", "测试步骤", "条件概述", "条件细节", "操作", "结果概述", "测试结果", "复测结果", "备注"
    ]

    # Set values, fill, and border for each cell in the specified row
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_num, header)
        cell.fill = heading_row_fill
        cell.border = border
        cell.font = bold_font
        cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')
        col_letter = string.ascii_uppercase[col_num - 1]
        ws.column_dimensions[col_letter].width = 20


    # ws.cell(current_row_number, 1, "被测模块")
    # ws.cell(current_row_number, 2, "用例名字")
    # ws.cell(current_row_number, 3, "测试步骤")
    # ws.cell(current_row_number, 4, "条件")
    # ws.cell(current_row_number, 5, "条件注释")
    # ws.cell(current_row_number, 6, "行动")
    # ws.cell(current_row_number, 7, "结果")
    # ws.cell(current_row_number, 8, "结果注释")
    # ws.cell(current_row_number, 9, "测试结果")
    # ws.cell(current_row_number, 10, "复测结果")
    # ws.cell(current_row_number, 11, "备注").fill = heading_row_fill
    # ws.cell(current_row_number, 11, "备注").border = border

    number_of_testcases = len(testset_json["test_cases"])
    test_case_start_row = 2

    for index_test_cases in range(number_of_testcases):

        # current_row_number = current_row_number + 1
        # test_case_start_row = current_row_number


        test_case_name = testset_json["test_cases"][index_test_cases]["test_case_name"]

        # module_name = "被测模块"
        # cel = ws.cell(current_row_number, 1, module_name)
        # set_cell_content(ws, current_row_number, 1, module_name)

        cell = ws.cell(test_case_start_row, 1, test_case_name)
        cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')
        # set_cell_content(ws, test_case_start_row, 2, test_case_name)
        

        number_of_test_steps = len(testset_json["test_cases"][index_test_cases]["test_steps"])
        step_row_number = 0
        for index_test_steps in range(number_of_test_steps):
            # current_row_number = current_row_number + index_test_steps
            step_row_number = test_case_start_row + index_test_steps

            condition = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["condition"]["condition_description"]
            number_of_condition_states = len(testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["condition"]["condition_states"])
        
            cell = ws.cell(step_row_number, 2, index_test_steps + 1) # 测试步骤: 2 - column2, test step = index_test_steps + 1
            # set_cell_content(ws, step_row_number, 2, index_test_steps + 1)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.border = border

            condition_content = condition # condition description
            condition_note = "" # condition details
 
            for index_condition_states in range(number_of_condition_states):
                variable = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["condition"]["condition_states"][index_condition_states]["variable_name"]

                state_name = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["condition"]["condition_states"][index_condition_states]["state_name"]
                # condition_content = condition_content + f"{variable}: {state_name}\n"

                number_of_attribute_equivalence_classes = len(testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["condition"]["condition_states"][index_condition_states]["attribute_equivalence_classes"])

                
                for index_attribute_equivalence_classes in range(number_of_attribute_equivalence_classes):
                    attribute_name = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["condition"]["condition_states"][index_condition_states]["attribute_equivalence_classes"][index_attribute_equivalence_classes]["attribute_name"]
                    attribute_description = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["condition"]["condition_states"][index_condition_states]["attribute_equivalence_classes"][index_attribute_equivalence_classes]["attribute_description"]
                    equivalence_class_name = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["condition"]["condition_states"][index_condition_states]["attribute_equivalence_classes"][index_attribute_equivalence_classes]["equivalence_class_name"]
                    equivalence_class_description = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["condition"]["condition_states"][index_condition_states]["attribute_equivalence_classes"][index_attribute_equivalence_classes]["equivalence_class_description"]

                    condition_note = condition_note + f"{variable}•{attribute_name}: {equivalence_class_name}\n"
            # ws.cell(current_row_number, 4, condition_content)
            set_cell_content(ws, step_row_number, 3, condition_content)
            # ws.cell(current_row_number, 5, condition_note)
            set_cell_content(ws, step_row_number, 4, condition_note[:-1])

            action_name = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["action_name"]
            # ws.cell(current_row_number, 6, action_name)
            set_cell_content(ws, step_row_number, 5, action_name)
            
            
            result = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["result"]["result_description"]
            number_of_result_states = len(testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["result"]["result_states"])

            result_content = result
            result_note = ""
            for index_result_states in range(number_of_result_states):
                variable = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["result"]["result_states"][index_result_states]["variable_name"]
                state_name = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["result"]["result_states"][index_result_states]["state_name"]
                # result_content = result_content + f"{variable}: {state_name}\n"

                number_of_attribute_equivalence_classes = len(testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["result"]["result_states"][index_result_states]["attribute_equivalence_classes"])

                for index_attribute_equivalence_classes in range(number_of_attribute_equivalence_classes):
                    attribute_name = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["result"]["result_states"][index_result_states]["attribute_equivalence_classes"][index_attribute_equivalence_classes]["attribute_name"]
                    attribute_description = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["result"]["result_states"][index_result_states]["attribute_equivalence_classes"][index_attribute_equivalence_classes]["attribute_description"]
                    equivalence_class_name = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["result"]["result_states"][index_result_states]["attribute_equivalence_classes"][index_attribute_equivalence_classes]["equivalence_class_name"]
                    equivalence_class_description = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["result"]["result_states"][index_result_states]["attribute_equivalence_classes"][index_attribute_equivalence_classes]["equivalence_class_description"]

                    result_note = result_note + f"{attribute_name}: {equivalence_class_name}\n"
            # ws.cell(current_row_number, 7, result_content)
            set_cell_content(ws, step_row_number, 6, result_content)
            # ws.cell(current_row_number, 8, result_note)
            # set_cell_content(ws, step_row_number, 7, result_note)
            set_cell_content(ws, step_row_number, 7, "")
            set_cell_content(ws, step_row_number, 8, "")
            set_cell_content(ws, step_row_number, 9, "")

            # current_row_number = current_row_number + 1

            # wb.save("/opt/bihua/reqgpt/seagent/output2.xlsx")
            # # mylist = list(ws.rows)
            # # print(mylist)
            # print(current_row_number)
            # print("filler")
        
        # ws.cell(index_test_cases, 2, test_case_name) # number_of_test_steps
        # merge test case name
        # current_row_number = current_row_number - 1
        module_start_cell = f'A{test_case_start_row}'  # A1 is the start cell (Column 1, Row 1) 测试用例
        module_end_cell = f'A{step_row_number}'  # End cell in Column 1, Row end_row
        ws.merge_cells(f'{module_start_cell}:{module_end_cell}')
        start_row, start_col = ws[module_start_cell].row, ws[module_start_cell].column
        end_row, end_col = ws[module_end_cell].row, ws[module_end_cell].column

        # Apply border to each cell in the merged range
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border

        # test_case_start_cell = f'B{test_case_start_row}'  # A1 is the start cell (Column 1, Row 1) 被测软件
        # test_case_end_cell = f'B{current_row_number}'  # End cell in Column 1, Row end_row
        # cell = ws.merge_cells(f'{test_case_start_cell}:{test_case_end_cell}')
        # start_row, start_col = ws[test_case_start_cell].row, ws[test_case_start_cell].column
        # end_row, end_col = ws[test_case_end_cell].row, ws[test_case_end_cell].column

        # Apply border to each cell in the merged range
        # for row in range(start_row, end_row + 1):
        #     for col in range(start_col, end_col + 1):
        #         cell = ws.cell(row=row, column=col)
        #         cell.border = border

        test_case_start_row = test_case_start_row + number_of_test_steps
        # wb.save("/opt/bihua/reqgpt/seagent/output2.xlsx")

    # wb.save("/opt/bihua/reqgpt/seagent/output2.xlsx")
    return wb

def generate_testset(user_identifier: str, project_name: string, testcase_scope:json):

    # purge gdb for this project (peoject name = project name + user_id), load specifications to gdb
    load_dotenv()
    project_path = oms_library_project.get_project_path(user_identifier, project_name)
    SPEC_SUB_DIRECTORY = os.getenv("SPEC_SUB_DIRECTORY")
    specification_directory = os.path.join(project_path, SPEC_SUB_DIRECTORY)
    project_name = st.session_state.app_state["project_name"]
    project_name_wirh_user_identifier = project_name + "_" + user_identifier
    project_name = oms_library_gdb.insert_all_oms_to_gdb(specification_directory, user_identifier, project_name_wirh_user_identifier)
    if project_name is None:
        return None
    # generate testset from gdb
    testset_json = generate_testset_json_from_gdb(specification_directory, project_name_wirh_user_identifier, testcase_scope)
    testset_xlsx_data = testset_convert_json_to_xlsx(testset_json)
    timestamp = time.time()
    # print("Excel完成（秒）:", timestamp)
    return testset_xlsx_data
