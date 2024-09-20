import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
import req_gdb
import seagent_file
import string

def generate_knowledge_base():
    pass

def generate_requirement():
    pass

def generate_test_set():
    # find all start cars
    # find all end cars
    # get all paths
    # save to test set template, for test automation
    # location = /opt/bihua/reqgpt/data/apps/eric/bookstore/test_cases


    return 

# Define border
border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

bold_font = Font(bold=True)
# Colors and borders setting:
odd_row_fill = PatternFill(start_color="FFFAEB", end_color="FFFAEB", fill_type="solid")
even_row_fill = PatternFill(start_color="FFF1C5", end_color="FFF1C5", fill_type="solid")
heading_row_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    # Apply styles to the cell
    # cell = ws['A1']
    # cell.fill = fill
    # cell.border = border

def set_cell_content(wb, row, column, value):
    cell = wb.cell(row, column, value)
    cell.border = border
    if row % 2 != 0: # odd
        cell.fill = odd_row_fill
    else:
        cell.fill = even_row_fill
    cell.alignment = Alignment(wrap_text=True)



def convert_json_to_xlsx(testset_json):
    path = "/opt/bihua/reqgpt/seagent/testcase_template.json"

    testset_json = seagent_file.oms_load(path)

    wb = Workbook()
    ws = wb.active
    ws.title = testset_json["system_under_test"]
    
    current_row_number = 1
    headers = [
    "被测模块", "用例名称", "测试步骤", "条件", "条件注释", "行动", "结果", "结果注释", "测试结果", "复测结果", "备注"
    ]

    # Set values, fill, and border for each cell in the specified row
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(current_row_number, col_num, header)
        cell.fill = heading_row_fill
        cell.border = border
        cell.font = bold_font
        cell.alignment = Alignment(wrap_text=True)
        col_letter = string.ascii_uppercase[col_num - 1]
        ws.column_dimensions[col_letter].width = 15


    # ws.cell(current_row_number, 1, "被测模块")
    # ws.cell(current_row_number, 2, "用例名字")
    # ws.cell(current_row_number, 3, "测试步骤")
    # ws.cell(current_row_number, 4, "条件")
    # ws.cell(current_row_number, 5, "条件注释")
    # ws.cell(current_row_number, 6, "行动")
    # ws.cell(current_row_number, 7, "结果")
    # ws.cell(current_row_number, 8, "结果注释")
    # ws.cell(current_row_number, 9, "测试结果")
    # ws.cell(current_row_number, 10, "复测结果")
    # ws.cell(current_row_number, 11, "备注").fill = heading_row_fill
    # ws.cell(current_row_number, 11, "备注").border = border

    number_of_testcases = len(testset_json["test_cases"])
    test_case_start_row = 0

    for index_test_cases in range(number_of_testcases):

        current_row_number = current_row_number + 1
        test_case_start_row = current_row_number


        test_case_name = testset_json["test_cases"][index_test_cases]["test_case_name"]

        module_name = "被测模块"
        cel = ws.cell(current_row_number, 1, module_name)
        # set_cell_content(ws, current_row_number, 1, module_name)

        ws.cell(current_row_number, 2, test_case_name)
        # set_cell_content(ws, current_row_number, 2, test_case_name)

        number_of_test_steps = len(testset_json["test_cases"][index_test_cases]["test_steps"])
        for index_test_steps in range(number_of_test_steps):
            current_row_number = current_row_number + index_test_steps

            # condition = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["condition"]["condition_description"]
            number_of_condition_states = len(testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["condition"]["condition_states"])
        
            ws.cell(current_row_number, 3, index_test_steps + 1) # 测试步骤
            set_cell_content(ws, current_row_number, 3, index_test_steps + 1)

            condition_content = ""
            condition_note = ""
            for index_condition_states in range(number_of_condition_states):
                variable = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["condition"]["condition_states"][index_condition_states]["variable_name"]
                state_name = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["condition"]["condition_states"][index_condition_states]["state_name"]
                condition_content = condition_content + f"{variable}: {state_name}\n"

                number_of_attribute_equivalence_classes = len(testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["condition"]["condition_states"][index_condition_states]["attribute_equivalence_classes"])

                
                for index_attribute_equivalence_classes in range(number_of_attribute_equivalence_classes):
                    attribute_name = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["condition"]["condition_states"][index_condition_states]["attribute_equivalence_classes"][index_attribute_equivalence_classes]["attribute_name"]
                    attribute_description = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["condition"]["condition_states"][index_condition_states]["attribute_equivalence_classes"][index_attribute_equivalence_classes]["attribute_description"]
                    equivalence_class_name = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["condition"]["condition_states"][index_condition_states]["attribute_equivalence_classes"][index_attribute_equivalence_classes]["equivalence_class_name"]
                    equivalence_class_description = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["condition"]["condition_states"][index_condition_states]["attribute_equivalence_classes"][index_attribute_equivalence_classes]["equivalence_class_description"]

                    condition_note = condition_note + f"{attribute_name}: {equivalence_class_name}\n"
            # ws.cell(current_row_number, 4, condition_content)
            set_cell_content(ws, current_row_number, 4, condition_content)
            # ws.cell(current_row_number, 5, condition_note)
            set_cell_content(ws, current_row_number, 5, condition_note)

            action_name = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["action_name"]
            # ws.cell(current_row_number, 6, action_name)
            set_cell_content(ws, current_row_number, 6, action_name)
            
            
            # result = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["result"]["result_description"]
            number_of_result_states = len(testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["result"]["result_states"])

            result_content = ""
            result_note = ""
            for index_result_states in range(number_of_result_states):
                variable = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["result"]["result_states"][index_result_states]["variable_name"]
                state_name = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["result"]["result_states"][index_result_states]["state_name"]
                result_content = result_content + f"{variable}: {state_name}\n"

                number_of_attribute_equivalence_classes = len(testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["result"]["result_states"][index_result_states]["attribute_equivalence_classes"])

                for index_attribute_equivalence_classes in range(number_of_attribute_equivalence_classes):
                    attribute_name = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["result"]["result_states"][index_result_states]["attribute_equivalence_classes"][index_attribute_equivalence_classes]["attribute_name"]
                    attribute_description = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["result"]["result_states"][index_result_states]["attribute_equivalence_classes"][index_attribute_equivalence_classes]["attribute_description"]
                    equivalence_class_name = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["result"]["result_states"][index_result_states]["attribute_equivalence_classes"][index_attribute_equivalence_classes]["equivalence_class_name"]
                    equivalence_class_description = testset_json["test_cases"][index_test_cases]["test_steps"][index_test_steps]["result"]["result_states"][index_result_states]["attribute_equivalence_classes"][index_attribute_equivalence_classes]["equivalence_class_description"]

                    result_note = result_note + f"{attribute_name}: {equivalence_class_name}\n"
            # ws.cell(current_row_number, 7, result_content)
            set_cell_content(ws, current_row_number, 7, result_content)
            # ws.cell(current_row_number, 8, result_note)
            set_cell_content(ws, current_row_number, 8, result_note)
            set_cell_content(ws, current_row_number, 9, "")
            set_cell_content(ws, current_row_number, 10, "")
            set_cell_content(ws, current_row_number, 11, "")
            
        
        # ws.cell(index_test_cases, 2, test_case_name) # number_of_test_steps

        module_start_cell = f'A{test_case_start_row}'  # A1 is the start cell (Column 1, Row 1) 被测软件
        module_end_cell = f'A{current_row_number}'  # End cell in Column 1, Row end_row
        ws.merge_cells(f'{module_start_cell}:{module_end_cell}')
        start_row, start_col = ws[module_start_cell].row, ws[module_start_cell].column
        end_row, end_col = ws[module_end_cell].row, ws[module_end_cell].column

        # Apply border to each cell in the merged range
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border

        test_case_start_cell = f'B{test_case_start_row}'  # A1 is the start cell (Column 1, Row 1) 被测软件
        test_case_end_cell = f'B{current_row_number}'  # End cell in Column 1, Row end_row
        cell = ws.merge_cells(f'{test_case_start_cell}:{test_case_end_cell}')
        start_row, start_col = ws[test_case_start_cell].row, ws[test_case_start_cell].column
        end_row, end_col = ws[test_case_end_cell].row, ws[test_case_end_cell].column

        # Apply border to each cell in the merged range
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border

        test_case_start_row = test_case_start_row + number_of_test_steps

    # wb.save("/opt/bihua/reqgpt/seagent/output2.xlsx")

# convert_json_to_xlsx("test")



    # ws.cell(row=index+1, column=i, value=value)
    # start_col_letter = get_column_letter(start_col_idx)
    # end_col_letter = get_column_letter(end_col_idx)
    # cell_range = f'{start_col_letter}{start_row}:{end_col_letter}{end_row}'
    # ws.merge_cells(cell_range)

    # # Hide column B (which is the 2nd column)
    # ws.column_dimensions['B'].hidden = True



    # process each test case, add cells one by one. merge the ones we want to merge.

