import json
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# Sample JSON data
json_data = '''
{
    "Sheet1": [
        {"A": "Header 1", "B": "Header 2", "C": "Header 3"},
        {"A": "Data 1", "B": "Data 2", "C": "Data 3"},
        {"A": "Data 4", "B": "Data 5", "C": "Data 6"}
    ],
    "MergedCells": [
        {"row": 1, "start_col": "A", "end_col": "C", "text": "Merged Header"}
    ]
}
'''

# Load JSON data
data = json.loads(json_data)

# Create a new workbook and get the active sheet
wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

# Populate sheet with data
for row in data["Sheet1"]:
    ws.append(row.values())

# Apply merged cells based on JSON instructions
for merge in data["MergedCells"]:
    ws.merge_cells(start_row=merge["row"], start_column=ord(merge["start_col"]) - 64,
                   end_row=merge["row"], end_column=ord(merge["end_col"]) - 64)
    cell = ws.cell(row=merge["row"], column=ord(merge["start_col"]) - 64)
    cell.value = merge["text"]

# Save the workbook
wb.save("output.xlsx")
