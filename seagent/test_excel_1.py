import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

# Sample HTML with merged cells
html_content = '''
<table border="1">
    <tr>
        <th rowspan="2">Header 1</th>
        <th colspan="2">Header 2</th>
    </tr>
    <tr>
        <th>Subheader 1</th>
        <th>Subheader 2</th>
    </tr>
    <tr>
        <td>Data 1</td>
        <td>Data 2</td>
        <td>Data 3</td>
    </tr>
</table>
'''

# Parse HTML content
soup = BeautifulSoup(html_content, 'html.parser')
table = soup.find('table')

# Extract table data
rows = []
for row in table.find_all('tr'):
    cells = [cell.get_text(strip=True) for cell in row.find_all(['td', 'th'])]
    rows.append(cells)

# Convert to DataFrame
df = pd.DataFrame(rows)

# Create an Excel workbook and add a worksheet
wb = Workbook()
ws = wb.active

# Write DataFrame to worksheet
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)

# Manually merge cells based on the HTML content
ws.merge_cells('A1:A2')  # Example: Merging Header 1 across rows
ws.merge_cells('B1:C1')  # Example: Merging Header 2 across columns

# Save the workbook
wb.save('output.xlsx')


wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

# Example data
ws['A1'] = "Styled Cell"

# Define background color
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side

fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Define border
border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

# Apply styles to the cell
cell = ws['A1']
cell.fill = fill
cell.border = border